"""إنشاء ملف Excel وصور PNG صغيرة لعرض التطبيق."""

from __future__ import annotations

import base64
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

PNG_1X1 = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


def create_samples(root: Path) -> None:
    """أنشئ العينات داخل مجلد samples."""

    sample_root = root / "samples"
    images = sample_root / "source_images"
    images.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "الأفراد"
    sheet.sheet_view.rightToLeft = True
    sheet.append(["الاسم", "الرقم الذاتي", "ملاحظات"])
    rows = [
        ("أحمد", "a3222263", "له صورة"),
        ("سارة", "000125", "رقم بأصفار بادئة"),
        ("محمد", "A-2026-15", "امتداد آخر"),
        ("ليلى", "غير_موجود", "للتجربة"),
        ("أحمد مكرر", "a3222263", "قيمة مكررة"),
        ("هند", "multi-77", "تطابقات متعددة"),
        ("سامر", "fuzzy-100", "اقتراح تقريبي"),
        ("ريم", "AR-٩٩", "اسم عربي وأرقام عربية"),
        ("نور", "PREFIX-42", "مثال بادئة ولاحقة"),
    ]
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(name="Tajawal", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17324D")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 28
    workbook.save(sample_root / "sample_identifiers.xlsx")
    for filename in (
        "a3222263.png",
        "000125.png",
        "A-2026-15.png",
        "multi-77.jpg",
        "multi-77.png",
        "fuzzy-10.png",
        "AR-٩٩.PNG",
        "IMG_PREFIX-42_front.JPG",
    ):
        (images / filename).write_bytes(PNG_1X1)
    nested = images / "مجلد_فرعي"
    nested.mkdir(exist_ok=True)
    (nested / "nested_001.webp").write_bytes(PNG_1X1)
    conflicts = sample_root / "destination_conflicts"
    conflicts.mkdir(exist_ok=True)
    (conflicts / "a3222263.png").write_bytes(PNG_1X1)
    (sample_root / "sample_identifiers.csv").write_text(
        "الاسم,الرقم الذاتي,ملاحظات\nأحمد,a3222263,مطابق\nسارة,000125,أصفار بادئة\nهند,multi-77,متعدد\nسامر,fuzzy-100,اقتراح\n",
        encoding="utf-8-sig",
    )


if __name__ == "__main__":
    create_samples(Path(__file__).resolve().parents[1])
