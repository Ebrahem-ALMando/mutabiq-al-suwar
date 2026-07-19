"""قراءة ملفات Excel واستخراج المعرّفات بدقة."""

from __future__ import annotations

import csv
import difflib
import math
import re
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter

from models.result_models import ColumnInfo, IdentifierRecord
from utils.constants import IDENTIFIER_HEADER_ALIASES, SUPPORTED_EXCEL_EXTENSIONS

_SPACE_RE = re.compile(r"\s+")
_TATWEEL_RE = re.compile("ـ+")
_ARABIC_MARKS_RE = re.compile(r"[\u0610-\u061A\u064B-\u065F\u0670\u06D6-\u06ED]")
_HEADER_PUNCTUATION_RE = re.compile(r"[\u200B-\u200F\u202A-\u202E:：،,؛;._\-]+")


def normalize_header(value: object) -> str:
    """طبّع عنوان عمود للمقارنة فقط مع إبقاء الحروف العربية كما هي."""

    if value is None:
        return ""
    text = _TATWEEL_RE.sub("", str(value))
    text = _ARABIC_MARKS_RE.sub("", text)
    text = _HEADER_PUNCTUATION_RE.sub(" ", text)
    return _SPACE_RE.sub(" ", text.replace("\r", " ").replace("\n", " ")).strip().casefold()


def _zero_padding_width(number_format: str) -> int | None:
    """استخرج عرض تنسيق أصفار صحيح وبسيط، مثل 000000."""

    primary = number_format.split(";", 1)[0].strip()
    primary = re.sub(r'"[^"]*"', "", primary)
    primary = re.sub(r"\[[^\]]*\]", "", primary)
    if re.fullmatch(r"0+", primary):
        return len(primary)
    return None


def cell_value_to_identifier(value: Any, number_format: str = "General") -> str:
    """حوّل قيمة خلية إلى معرّف من دون إضافة .0 وباحترام تنسيق الأصفار."""

    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        width = _zero_padding_width(number_format)
        return f"{value:0{width}d}" if width and value >= 0 else str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return str(value)
        if value.is_integer():
            integer = int(value)
            width = _zero_padding_width(number_format)
            return f"{integer:0{width}d}" if width and integer >= 0 else str(integer)
        return format(Decimal(str(value)).normalize(), "f")
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            integer = int(value)
            width = _zero_padding_width(number_format)
            return f"{integer:0{width}d}" if width and integer >= 0 else str(integer)
        return format(value.normalize(), "f")
    return str(value)


@dataclass(slots=True)
class WorkbookInfo:
    """أسماء الأوراق والورقة النشطة."""

    worksheets: list[str]
    active_worksheet: str


@dataclass(slots=True)
class ColumnScore:
    """درجة ترشيح عمود مع تفسير وثقة قابلة للعرض."""

    column: ColumnInfo
    score: float
    confidence: str
    reasons: list[str]


@dataclass(slots=True)
class DelimitedFileInfo:
    """نتيجة اكتشاف ترميز وفاصل ملف نصي."""

    encoding: str
    delimiter: str
    confident: bool


def detect_delimited_format(path: Path) -> DelimitedFileInfo:
    """اكتشف ترميزاً عربياً شائعاً وفاصلاً مع fallback حتمي."""

    raw = path.read_bytes()[:131072]
    decoded = ""
    encoding = "utf-8"
    for candidate in ("utf-8-sig", "utf-8", "cp1256", "iso-8859-6", "cp1252"):
        try:
            decoded = raw.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    else:
        decoded = raw.decode("utf-8", errors="replace")
    default = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(decoded, delimiters=",;\t|")
        return DelimitedFileInfo(encoding, dialect.delimiter, True)
    except csv.Error:
        return DelimitedFileInfo(encoding, default, False)


def _header_row(values_by_row: list[list[Any]]) -> tuple[int, list[Any]]:
    """اختر صف العناوين الأفضل من عينة صفوف."""

    best_row = 1
    best_values: list[Any] = []
    aliases = {normalize_header(alias) for alias in IDENTIFIER_HEADER_ALIASES}
    for row_index, values in enumerate(values_by_row, start=1):
        non_empty = sum(bool(value is not None and str(value).strip()) for value in values)
        if any(normalize_header(value) in aliases for value in values):
            return row_index, values
        previous_count = sum(bool(value is not None and str(value).strip()) for value in best_values)
        if non_empty > previous_count:
            best_row, best_values = row_index, values
    return best_row, best_values


class ExcelService:
    """واجهة قابلة للاختبار لجميع عمليات قراءة Excel."""

    def validate_path(self, path: Path) -> None:
        """تحقق من وجود ملف Excel وامتداده."""

        if not path.is_file():
            raise FileNotFoundError("ملف Excel غير موجود.")
        if path.suffix.lower() not in SUPPORTED_EXCEL_EXTENSIONS:
            raise ValueError("صيغة ملف Excel غير مدعومة. استخدم xlsx أو xlsm.")

    def workbook_info(self, path: Path) -> WorkbookInfo:
        """اقرأ أسماء الأوراق واسم الورقة النشطة دون تعديل الملف."""

        self.validate_path(path)
        if path.suffix.lower() in {".csv", ".tsv"}:
            return WorkbookInfo(["البيانات"], "البيانات")
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            names = list(workbook.sheetnames)
            if not names:
                raise ValueError("لا توجد أوراق عمل في الملف.")
            return WorkbookInfo(names, workbook.active.title)
        finally:
            workbook.close()

    def columns(self, path: Path, worksheet: str, search_rows: int = 25) -> list[ColumnInfo]:
        """اكتشف صف العناوين وأعد الأعمدة غير الفارغة."""

        self.validate_path(path)
        if path.suffix.lower() in {".csv", ".tsv"}:
            info = detect_delimited_format(path)
            with path.open("r", encoding=info.encoding, newline="") as stream:
                rows = []
                reader = csv.reader(stream, delimiter=info.delimiter)
                for _, row in zip(range(search_rows), reader, strict=False):
                    rows.append(list(row))
            best_row, best_values = _header_row(rows)
            return [
                ColumnInfo(index=index, header=str(value).strip(), letter=get_column_letter(index), header_row=best_row)
                for index, value in enumerate(best_values, start=1)
                if value is not None and str(value).strip()
            ]
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if worksheet not in workbook.sheetnames:
                raise ValueError("ورقة العمل المحددة غير موجودة.")
            sheet = workbook[worksheet]
            rows = [list(row) for row in sheet.iter_rows(max_row=search_rows, values_only=True)]
            best_row, best_values = _header_row(rows)
            return [
                ColumnInfo(index=index, header=str(value).strip(), letter=get_column_letter(index), header_row=best_row)
                for index, value in enumerate(best_values, start=1)
                if value is not None and str(value).strip()
            ]
        finally:
            workbook.close()

    @staticmethod
    def find_required_column(columns: list[ColumnInfo]) -> ColumnInfo | None:
        """ابحث عن عمود الرقم الذاتي بمطابقة عنوان متسامحة."""

        aliases = {normalize_header(alias) for alias in IDENTIFIER_HEADER_ALIASES}
        return next((column for column in columns if normalize_header(column.header) in aliases), None)

    def read_identifiers(
        self,
        path: Path,
        worksheet: str,
        column: ColumnInfo,
        trim: bool = True,
        case_insensitive: bool = True,
        start_row: int | None = None,
        end_row: int | None = None,
        secondary_column: ColumnInfo | None = None,
    ) -> list[IdentifierRecord]:
        """استخرج جميع صفوف العمود المحدد مع اكتشاف التكرارات."""

        self.validate_path(path)
        if path.suffix.lower() in {".csv", ".tsv"}:
            return self._read_delimited_identifiers(
                path, column, trim, case_insensitive, start_row, end_row, secondary_column
            )
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if worksheet not in workbook.sheetnames:
                raise ValueError("ورقة العمل المحددة غير موجودة.")
            sheet = workbook[worksheet]
            records: list[IdentifierRecord] = []
            seen: set[str] = set()
            first = max(column.header_row + 1, start_row or column.header_row + 1)
            last = min(sheet.max_row, end_row or sheet.max_row)
            for row_number in range(first, last + 1):
                cell: Cell = sheet.cell(row=row_number, column=column.index)
                identifier = cell_value_to_identifier(cell.value, cell.number_format)
                if trim:
                    identifier = identifier.strip()
                empty = identifier == ""
                key = identifier.casefold() if case_insensitive else identifier
                duplicate = not empty and key in seen
                if not empty:
                    seen.add(key)
                records.append(
                    IdentifierRecord(
                        row_number=row_number,
                        original_value=cell.value,
                        identifier=identifier,
                        is_empty=empty,
                        is_duplicate=duplicate,
                        secondary_name=(
                            cell_value_to_identifier(sheet.cell(row=row_number, column=secondary_column.index).value)
                            if secondary_column
                            else ""
                        ),
                    )
                )
            return records
        finally:
            workbook.close()

    def _read_delimited_identifiers(
        self,
        path: Path,
        column: ColumnInfo,
        trim: bool,
        case_insensitive: bool,
        start_row: int | None,
        end_row: int | None,
        secondary_column: ColumnInfo | None,
    ) -> list[IdentifierRecord]:
        info = detect_delimited_format(path)
        records: list[IdentifierRecord] = []
        seen: set[str] = set()
        first = max(column.header_row + 1, start_row or column.header_row + 1)
        with path.open("r", encoding=info.encoding, newline="") as stream:
            for row_number, row in enumerate(csv.reader(stream, delimiter=info.delimiter), start=1):
                if row_number < first:
                    continue
                if end_row and row_number > end_row:
                    break
                value = row[column.index - 1] if len(row) >= column.index else ""
                identifier = value.strip() if trim else value
                empty = identifier == ""
                key = identifier.casefold() if case_insensitive else identifier
                duplicate = not empty and key in seen
                if not empty:
                    seen.add(key)
                secondary = ""
                if secondary_column and len(row) >= secondary_column.index:
                    secondary = row[secondary_column.index - 1]
                records.append(IdentifierRecord(row_number, value, identifier, empty, duplicate, secondary))
        return records

    def preview_rows(self, path: Path, worksheet: str, limit: int = 20) -> list[list[str]]:
        """أعد عينة بيانات أولية لاستخدامها في شاشة الاستيراد."""

        self.validate_path(path)
        if path.suffix.lower() in {".csv", ".tsv"}:
            info = detect_delimited_format(path)
            with path.open("r", encoding=info.encoding, newline="") as stream:
                return [
                    [str(value) for value in row]
                    for _, row in zip(range(limit), csv.reader(stream, delimiter=info.delimiter), strict=False)
                ]
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[worksheet]
            return [
                [cell_value_to_identifier(value) for value in row]
                for row in sheet.iter_rows(max_row=limit, values_only=True)
            ]
        finally:
            workbook.close()

    def score_columns(self, path: Path, worksheet: str, columns: list[ColumnInfo]) -> list[ColumnScore]:
        """رتّب الأعمدة بحسب الاسم والاكتمال والتفرّد واتساق القيم."""

        preview = self.preview_rows(path, worksheet, 250)
        aliases = [normalize_header(alias) for alias in IDENTIFIER_HEADER_ALIASES]
        scored: list[ColumnScore] = []
        for column in columns:
            header = normalize_header(column.header)
            name_similarity = max(difflib.SequenceMatcher(None, header, alias).ratio() for alias in aliases)
            values = [row[column.index - 1].strip() for row in preview[column.header_row :] if len(row) >= column.index]
            non_empty = [value for value in values if value]
            completeness = len(non_empty) / max(1, len(values))
            uniqueness = len(set(non_empty)) / max(1, len(non_empty))
            patterned = sum(bool(re.fullmatch(r"[\w\- ]+", value, re.UNICODE)) for value in non_empty) / max(
                1, len(non_empty)
            )
            score = 60 * name_similarity + 20 * completeness + 15 * uniqueness + 5 * patterned
            reasons = [f"تشابه العنوان {name_similarity:.0%}", f"اكتمال {completeness:.0%}", f"تفرّد {uniqueness:.0%}"]
            confidence = "ثقة عالية" if score >= 78 else "ثقة متوسطة" if score >= 55 else "ثقة منخفضة"
            scored.append(ColumnScore(column, round(score, 1), confidence, reasons))
        return sorted(scored, key=lambda item: (-item.score, item.column.index))
