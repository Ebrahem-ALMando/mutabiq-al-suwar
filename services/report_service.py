"""إنشاء تقرير Excel منسق وكامل لنتائج المعالجة."""

from __future__ import annotations

import csv
import html
import os
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from models.result_models import JobResult, MatchStatus
from utils.constants import COPY_STATUS_AR, DUPLICATE_POLICY_AR, MATCH_STATUS_AR, MULTIPLE_POLICY_AR
from utils.path_helpers import unique_destination
from utils.version import APP_VERSION


class ReportService:
    """يبني ملف Excel لا يعتمد على وجود Microsoft Excel."""

    HEADER_FILL = PatternFill("solid", fgColor="17324D")
    ACCENT_FILL = PatternFill("solid", fgColor="E8F1F8")
    SUCCESS_FILL = PatternFill("solid", fgColor="DCFCE7")
    WARNING_FILL = PatternFill("solid", fgColor="FEF3C7")
    ERROR_FILL = PatternFill("solid", fgColor="FEE2E2")
    THIN_BORDER = Border(bottom=Side(style="thin", color="CBD5E1"))

    def create(self, result: JobResult) -> Path:
        """أنشئ تقريراً بخمس أوراق داخل مجلد الوجهة."""

        result.settings.destination_folder.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = unique_destination(result.settings.destination_folder / f"تقرير_نسخ_الصور_{timestamp}.xlsx")
        workbook = Workbook()
        summary = workbook.active
        summary.title = "الملخص"
        self._write_summary(summary, result)
        self._write_records(workbook.create_sheet("جميع النتائج"), result.records, "AllResults")
        self._write_records(
            workbook.create_sheet("غير الموجودة"),
            [record for record in result.records if record.match_status == MatchStatus.NOT_FOUND],
            "NotFound",
        )
        self._write_records(
            workbook.create_sheet("الأخطاء"),
            [
                record
                for record in result.records
                if record.match_status == MatchStatus.ERROR or record.copy_status.value == "failed"
            ],
            "Errors",
        )
        self._write_records(
            workbook.create_sheet("التكرارات"),
            [
                record
                for record in result.records
                if record.match_status in {MatchStatus.DUPLICATE, MatchStatus.MULTIPLE}
            ],
            "Duplicates",
        )
        for sheet in workbook.worksheets:
            sheet.sheet_view.rightToLeft = True
        temporary_path = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
        try:
            workbook.save(temporary_path)
            os.replace(temporary_path, path)
        finally:
            try:
                temporary_path.unlink(missing_ok=True)
            except OSError:
                pass
        return path

    def create_csv(self, result: JobResult, path: Path | None = None) -> Path:
        """صدّر جميع النتائج إلى CSV UTF-8 مناسب لـ Excel."""

        path = path or result.settings.destination_folder / f"results_{result.batch_id}.csv"
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.writer(stream)
            writer.writerow(
                ["التسلسل", "صف Excel", "المعرّف", "الاسم", "حالة المطابقة", "المصدر", "الوجهة", "حالة النسخ", "ملاحظات"]
            )
            for record in result.records:
                writer.writerow(
                    [
                        record.sequence,
                        record.excel_row,
                        record.identifier,
                        record.secondary_name,
                        MATCH_STATUS_AR[record.match_status.value],
                        str(record.source_path or ""),
                        str(record.destination_path or ""),
                        COPY_STATUS_AR[record.copy_status.value],
                        record.notes,
                    ]
                )
        return path

    def create_html(self, result: JobResult, path: Path | None = None) -> Path:
        """أنشئ تقرير HTML عربي قابلاً للطباعة دون تبعيات إضافية."""

        path = path or result.settings.destination_folder / f"report_{result.batch_id}.html"
        rows = "".join(
            "<tr>"
            + "".join(
                f"<td>{html.escape(str(value))}</td>"
                for value in (
                    record.sequence,
                    record.identifier,
                    record.secondary_name,
                    MATCH_STATUS_AR[record.match_status.value],
                    record.source_filename,
                    record.destination_filename,
                    COPY_STATUS_AR[record.copy_status.value],
                    record.notes,
                )
            )
            + "</tr>"
            for record in result.records
        )
        document = f"""<!doctype html><html lang='ar' dir='rtl'><head><meta charset='utf-8'>
        <title>تقرير مُطابق الصور</title><style>body{{font-family:Tajawal,Segoe UI,Arial;margin:32px;color:#17324d}}
        h1{{color:#0f6b78}}table{{border-collapse:collapse;width:100%}}th,td{{border:1px solid #d8e2e8;padding:8px;text-align:right}}
        th{{background:#17324d;color:white}}tr:nth-child(even){{background:#f4f7fa}}.meta{{background:#edf5f6;padding:16px;border-radius:10px}}</style></head>
        <body><h1>تقرير مُطابق الصور</h1><div class='meta'>معرّف الدفعة: {html.escape(result.batch_id)}<br>
        ملف البيانات: {html.escape(str(result.settings.excel_path))}<br>النتيجة: {html.escape(result.outcome)}</div>
        <h2>النتائج</h2><table><thead><tr><th>#</th><th>المعرّف</th><th>الاسم</th><th>المطابقة</th><th>المصدر</th><th>الوجهة</th><th>النسخ</th><th>ملاحظات</th></tr></thead><tbody>{rows}</tbody></table></body></html>"""
        path.write_text(document, encoding="utf-8")
        return path

    def _write_summary(self, sheet, result: JobResult) -> None:
        sheet.merge_cells("A1:D1")
        sheet["A1"] = "تقرير نسخ الصور حسب الرقم الذاتي"
        sheet["A1"].font = Font(name="Tajawal", bold=True, size=18, color="FFFFFF")
        sheet["A1"].fill = self.HEADER_FILL
        sheet["A1"].alignment = Alignment(horizontal="center")
        rows = [
            ("تاريخ المعالجة", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("إصدار التطبيق", APP_VERSION),
            ("معرّف الدفعة", result.batch_id),
            ("محاكاة دون نسخ", "نعم" if result.settings.dry_run else "لا"),
            ("ملف Excel", str(result.settings.excel_path)),
            ("ورقة العمل", result.settings.worksheet),
            ("العمود", result.settings.column.header),
            ("مجلد الصور", str(result.settings.source_folder)),
            ("مجلد الحفظ", str(result.settings.destination_folder)),
            ("بحث داخل المجلدات الفرعية", "نعم" if result.settings.recursive else "لا"),
            ("مطابقة غير حساسة لحالة الأحرف", "نعم" if result.settings.case_insensitive else "لا"),
            ("سياسة الملفات الموجودة", DUPLICATE_POLICY_AR[result.settings.duplicate_policy.value]),
            ("سياسة التطابقات المتعددة", MULTIPLE_POLICY_AR[result.settings.multiple_match_policy.value]),
            ("نمط المطابقة", result.settings.matching_mode.value),
            (
                "قواعد التحويل",
                "; ".join(rule.kind for rule in result.settings.transformation_rules if rule.enabled) or "لا توجد",
            ),
            ("إجمالي صفوف Excel", result.stats.total_excel_rows),
            ("القيم الفارغة", result.stats.empty_values),
            ("المعرّفات الصالحة", result.stats.valid_identifiers),
            ("المعرّفات الفريدة", result.stats.unique_identifiers),
            ("القيم المكررة", result.stats.duplicate_identifiers),
            ("الصور المفهرسة", result.stats.source_images_scanned),
            ("المعرّفات المطابقة", result.stats.matched_identifiers),
            ("المعرّفات غير الموجودة", result.stats.unmatched_identifiers),
            ("التطابقات المتعددة", result.stats.multiple_match_identifiers),
            ("الملفات المنسوخة", result.stats.copied_files),
            ("الملفات المتخطاة", result.stats.skipped_files),
            ("الملفات المعاد تسميتها", result.stats.renamed_files),
            ("عمليات النسخ الفاشلة", result.stats.failed_copies),
            ("المدة بالثواني", round(result.stats.elapsed_seconds, 2)),
        ]
        for row_index, (label, value) in enumerate(rows, start=3):
            sheet.cell(row_index, 1, label)
            sheet.cell(row_index, 2, value)
            sheet.cell(row_index, 1).font = Font(name="Tajawal", bold=True, color="17324D")
            sheet.cell(row_index, 1).fill = self.ACCENT_FILL
            for column in range(1, 3):
                cell = sheet.cell(row_index, column)
                cell.font = Font(name="Tajawal", bold=column == 1)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = self.THIN_BORDER
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 80

    def _write_records(self, sheet, records, table_name: str) -> None:
        headers = [
            "التسلسل",
            "رقم صف Excel",
            "القيمة الأصلية",
            "الرقم المعالج",
            "حالة المطابقة",
            "اسم ملف المصدر",
            "مسار المصدر",
            "اسم ملف الوجهة",
            "مسار الوجهة",
            "حالة النسخ",
            "ملاحظات",
        ]
        sheet.append(headers)
        for record in records:
            sheet.append(
                [
                    record.sequence,
                    record.excel_row,
                    record.original_value,
                    record.identifier,
                    MATCH_STATUS_AR[record.match_status.value],
                    record.source_filename,
                    str(record.source_path or ""),
                    record.destination_filename,
                    str(record.destination_path or ""),
                    COPY_STATUS_AR[record.copy_status.value],
                    record.notes,
                ]
            )
        for cell in sheet[1]:
            cell.font = Font(name="Tajawal", bold=True, color="FFFFFF")
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Tajawal", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = self.THIN_BORDER
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = [11, 15, 22, 24, 20, 28, 55, 28, 55, 22, 45]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        if records:
            table = Table(displayName=table_name, ref=f"A1:K{len(records) + 1}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
            sheet.add_table(table)
            end_row = len(records) + 1
            sheet.conditional_formatting.add(
                f"E2:E{end_row}", FormulaRule(formula=['E2="مطابق"'], fill=self.SUCCESS_FILL)
            )
            sheet.conditional_formatting.add(
                f"E2:E{end_row}", FormulaRule(formula=['E2="غير موجود"'], fill=self.ERROR_FILL)
            )
            sheet.conditional_formatting.add(
                f"E2:E{end_row}", FormulaRule(formula=['E2="أكثر من تطابق"'], fill=self.WARNING_FILL)
            )
