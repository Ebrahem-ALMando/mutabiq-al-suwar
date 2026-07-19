"""اختبارات قراءة Excel وتحويل المعرّفات."""

from decimal import Decimal

from openpyxl import Workbook

from models.result_models import ColumnInfo
from services.excel_service import ExcelService, cell_value_to_identifier, normalize_header


def test_text_identifier_is_preserved() -> None:
    assert cell_value_to_identifier("A-2026_ 15") == "A-2026_ 15"


def test_integer_identifier() -> None:
    assert cell_value_to_identifier(123456) == "123456"


def test_integral_float_has_no_dot_zero() -> None:
    assert cell_value_to_identifier(123.0) == "123"


def test_decimal_identifier_is_not_rounded() -> None:
    assert cell_value_to_identifier(Decimal("12.340")) == "12.34"


def test_numeric_leading_zeros_from_format() -> None:
    assert cell_value_to_identifier(125, "000000") == "000125"


def test_header_normalization_tolerates_spaces_tatweel_and_lines() -> None:
    assert normalize_header("  الرقـم\n  الذاتي  ") == normalize_header("الرقم الذاتي")


def test_header_normalization_tolerates_diacritics_and_punctuation() -> None:
    assert normalize_header("الرَّقم: الذاتي") == normalize_header("الرقم الذاتي")


def test_reads_identifiers_and_finds_duplicates(tmp_path) -> None:
    path = tmp_path / "data.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "بيانات"
    sheet.append(["الاسم", "الرقم الذاتي"])
    sheet.append(["أ", "ABC"])
    sheet.append(["ب", "abc"])
    sheet.append(["ج", None])
    cell = sheet.cell(5, 2, 125)
    cell.number_format = "000000"
    workbook.save(path)

    service = ExcelService()
    columns = service.columns(path, "بيانات")
    selected = service.find_required_column(columns)
    assert selected is not None
    records = service.read_identifiers(path, "بيانات", selected, case_insensitive=True)
    assert [record.identifier for record in records] == ["ABC", "abc", "", "000125"]
    assert records[1].is_duplicate
    assert records[2].is_empty


def test_discovers_header_below_first_row(tmp_path) -> None:
    path = tmp_path / "offset.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["عنوان التقرير"])
    sheet.append([])
    sheet.append(["الرقم الذاتي", "ملاحظة"])
    workbook.save(path)
    columns = ExcelService().columns(path, sheet.title)
    assert columns[0] == ColumnInfo(1, "الرقم الذاتي", "A", 3)
